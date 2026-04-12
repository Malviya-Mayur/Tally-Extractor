"""
pipeline_runner.py — Background thread executor for the Tally Pipeline V2.

Imports functions from Tally_Pipeline_V2.py (one level up) and runs them
in a dedicated thread so the FastAPI server is never blocked.

Supports two source modes:
  1. Live Tally API  — fetches XML in monthly chunks from the running Tally instance.
  2. XML file upload — reads a previously-saved XML file (params["xml_file"]).

Output is always a single .xlsx workbook with two sheets:
  • "Data"           — flat transaction dump rows
  • "Extraction Log" — timestamped log lines for the job

V2 key changes vs V1 runner:
  - Imports Tally_Pipeline_V2 instead of Tally_Pipeline
  - Live mode fetches in monthly chunks via _month_chunks() — bounded RAM
  - Rows written to a temp SQLite DB per chunk via process_chunk_to_db()
  - CSV exported from SQLite via export_db_to_csv() (streaming, no full list in RAM)
  - Dimension exports come from MasterData.dimensions (merged across chunks)
  - chunk_months param added (default 1)
  - Default timeout raised to 300s
"""

from __future__ import annotations

import csv
import logging
import sqlite3
import sys
import tempfile
import threading
from pathlib import Path
from typing import Any, Callable

# ── Make the parent directory importable so we can reach Tally_Pipeline_V2 ───
_BASE = Path(__file__).resolve().parent.parent.parent  # …/Tally-Extractor/
if str(_BASE) not in sys.path:
    sys.path.insert(0, str(_BASE))

import Tally_Pipeline_V2 as tp  # noqa: E402  (after sys.path fix)

from . import jobs  # noqa: E402

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

class _JobLogHandler(logging.Handler):
    """A logging.Handler that forwards records to the in-memory job log."""

    def __init__(self, job_id: str, extra_callback: Callable[[str], None] | None = None):
        super().__init__()
        self.job_id = job_id
        self.extra_callback = extra_callback

    def emit(self, record: logging.LogRecord) -> None:
        line = self.format(record)
        jobs.append_log(self.job_id, line)
        if self.extra_callback:
            self.extra_callback(line)


def _write_xlsx(xlsx_path: Path, flat_rows: list[dict], log_lines: list[str]) -> None:
    """Write a two-sheet Excel workbook: 'Data' and 'Extraction Log'."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Data ─────────────────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Data"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    if flat_rows:
        headers = list(flat_rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(flat_rows, 2):
            for col_idx, key in enumerate(headers, 1):
                ws_data.cell(row=row_idx, column=col_idx, value=row.get(key))

        # Auto-fit column widths (cap at 50)
        for col_idx, header in enumerate(headers, 1):
            max_len = max(
                len(str(header)),
                *(len(str(row.get(header, "") or "")) for row in flat_rows[:200]),
            )
            ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    else:
        ws_data.cell(row=1, column=1, value="No data rows were produced.")

    # ── Sheet 2: Extraction Log ───────────────────────────────────────────────
    ws_log = wb.create_sheet("Extraction Log")
    log_header_fill = PatternFill("solid", fgColor="2D2D2D")
    log_header_font = Font(bold=True, color="FFFFFF")

    hdr = ws_log.cell(row=1, column=1, value="Extraction Log")
    hdr.fill = log_header_fill
    hdr.font = log_header_font
    ws_log.column_dimensions["A"].width = 120

    for row_idx, line in enumerate(log_lines, 2):
        ws_log.cell(row=row_idx, column=1, value=line)

    wb.save(xlsx_path)


def _read_csv_as_dicts(csv_path: Path) -> list[dict]:
    """Read a CSV file into a list of dicts (for the Excel writer)."""
    rows: list[dict] = []
    try:
        with csv_path.open(newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
    except Exception as exc:
        logger.warning("Could not read CSV %s: %s", csv_path, exc)
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def run_extraction(job_id: str, params: dict[str, Any]) -> None:
    """
    Execute the full Tally Pipeline V2 for a given job_id using params dict.

    params keys:
      from_date       str  YYYYMMDD  (required for live mode)
      to_date         str  YYYYMMDD  (required for live mode)
      port            int  Tally HTTP port (default 9000, live mode only)
      out_dir         str  Output directory path
      retries         int  Max HTTP retry attempts (live mode only)
      timeout         int  HTTP timeout in seconds (live mode only, default 300)
      chunk_months    int  Months per Tally HTTP request (default 1, live mode only)
      export_dims     list[str]  Dimension table names to export (e.g. ["LEDGER"])
      export_facts    list[str]  Fact table names: "voucher", "ledger_entry", "inventory_line"
      xml_file        str  (optional) Path to an already-extracted XML file. When
                           provided, the live Tally API call is skipped entirely.
    """
    import xml.etree.ElementTree as ET
    import re
    from datetime import datetime as _dt

    jobs.set_status(job_id, jobs.RUNNING)

    out_dir = Path(params.get("out_dir", "tally_out"))
    out_dir.mkdir(parents=True, exist_ok=True)

    # In-memory log collector for the Excel sheet
    collected_log_lines: list[str] = []

    # Set up a dedicated log handler that writes to the job store
    handler = _JobLogHandler(job_id)
    handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

    tp_logger = logging.getLogger("__main__")
    root_logger = logging.getLogger()
    root_logger.addHandler(handler)
    tp_logger.addHandler(handler)

    # Temp paths (cleaned up in finally)
    db_path: Path | None = None
    tmp_csv_path: Path | None = None
    xml_file: str | None = params.get("xml_file")

    def log(msg: str) -> None:
        jobs.append_log(job_id, msg)
        collected_log_lines.append(msg)

    try:
        export_dims: list[str] = params.get("export_dims", [])
        export_facts: list[str] = params.get("export_facts", [])

        # Temp SQLite DB path
        tmp_dir = Path(tempfile.mkdtemp(prefix="tally_v2_"))
        db_path = tmp_dir / "tally_pipeline_v2.db"

        # ── Initialise shared state across chunks ─────────────────────────────
        db_conn: sqlite3.Connection = sqlite3.connect(str(db_path))
        tp._init_db(db_conn)

        masters_global: tp.MasterData | None = None
        total_vouchers = 0
        total_rows = 0
        company_name = "company"
        from_date: str = params.get("from_date", "")
        to_date: str = params.get("to_date", "")

        # ── Build chunk list ──────────────────────────────────────────────────
        if xml_file:
            # XML Upload mode: treat the file as a single "chunk"
            chunk_list = [("xml_upload", "xml_upload")]
        else:
            chunk_months: int = int(params.get("chunk_months", 1))
            port: int = int(params.get("port", 9000))
            retries: int = int(params.get("retries", 3))
            timeout: int = int(params.get("timeout", 300))
            chunk_list = tp._month_chunks(from_date, to_date, chunk_months)
            log(
                f"Date range {from_date} → {to_date} split into "
                f"{len(chunk_list)} chunk(s) of {chunk_months} month(s) each."
            )

        # ── Main chunk loop ───────────────────────────────────────────────────
        for chunk_idx, (chunk_from, chunk_to) in enumerate(chunk_list):
            log(f"Processing chunk {chunk_idx + 1}/{len(chunk_list)} …")

            # Step 1: Obtain raw XML bytes
            if xml_file:
                xml_path = Path(xml_file)
                log(f"Reading XML from uploaded file: {xml_path.name} …")
                raw_bytes = xml_path.read_bytes()
                log(f"Read {len(raw_bytes):,} bytes from file.")
            else:
                log(f"Fetching XML from Tally at port {port} for {chunk_from} → {chunk_to} …")
                raw_bytes = tp.fetch_tally_xml_bytes(
                    chunk_from, chunk_to,
                    port=port,
                    max_retries=retries,
                    timeout=timeout,
                )
                if raw_bytes is None:
                    log(f"⚠ Chunk {chunk_idx + 1} failed to fetch — skipping.")
                    continue
                log(f"Received {len(raw_bytes):,} bytes.")

            # Step 2: Parse XML
            log("Parsing XML …")
            try:
                root = tp.parse_xml_from_bytes(raw_bytes)
            except ET.ParseError as exc:
                raise RuntimeError(f"Malformed XML in chunk {chunk_idx + 1}: {exc}") from exc

            # Extract company name from first successful chunk
            if company_name == "company":
                company_name = tp._safe_filename(tp._extract_company_name(root))

            # Step 3: Collect all masters (single tree-walk)
            log("Collecting masters and building element index …")
            masters = tp.collect_all_masters(root)

            # Merge into global masters so cross-chunk ledger lookups work
            if masters_global is None:
                masters_global = masters
            else:
                masters_global.ledger.update(masters.ledger)
                masters_global.stock.update(masters.stock)
                masters_global.group_parent.update(masters.group_parent)
                masters_global.stock_parent.update(masters.stock_parent)
                masters_global.voucher_sidecar.update(masters.voucher_sidecar)
                masters_global.credit_period.update(masters.credit_period)
                masters_global.le_elem.update(masters.le_elem)
                masters_global.inv_elem.update(masters.inv_elem)
                masters_global.party_le_elem.update(masters.party_le_elem)
                for dim_key, dim_rows in masters.dimensions.items():
                    masters_global.dimensions.setdefault(dim_key, []).extend(dim_rows)

            # Check for empty result
            has_vouchers = any(
                tp.strip_ns(c.tag) == "VOUCHER"
                for tm in root.iter()
                if tp.strip_ns(tm.tag) == "TALLYMESSAGE"
                for c in tm
            )
            if not has_vouchers:
                log(
                    f"⚠ Chunk {chunk_idx + 1}/{len(chunk_list)} returned zero vouchers. "
                    "Verify the TDL report 'APIRawVouchers' is loaded in Tally (F4 to load)."
                )
                del root, masters, raw_bytes
                continue

            # Step 4: Write flat fact rows directly to SQLite
            log("Writing flat fact rows to SQLite …")
            v_count, r_count = tp.process_chunk_to_db(root, masters, db_conn)
            total_vouchers += v_count
            total_rows += r_count
            log(
                f"Chunk {chunk_idx + 1}/{len(chunk_list)}: "
                f"{v_count:,} vouchers | {r_count:,} rows inserted."
            )

            # Release parsed tree before next chunk (free RAM)
            del root, masters, raw_bytes

        db_conn.close()

        if masters_global is None:
            raise RuntimeError(
                "No data was retrieved from Tally. "
                "Check that Tally is running and the TDL report 'APIRawVouchers' is loaded."
            )

        log(f"Total: {total_vouchers:,} vouchers | {total_rows:,} rows in SQLite.")

        # ── Step 5: Export sqlite → flat CSV (streaming, no full list in RAM) ──
        timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
        period = tp._format_period(from_date, to_date) if from_date and to_date else timestamp
        base_name = f"{company_name}_extraction_{period}_{timestamp}"

        tmp_csv_path = tmp_dir / f"{base_name}.csv"
        log("Exporting rows from SQLite to CSV …")
        csv_row_count = tp.export_db_to_csv(db_path, tmp_csv_path)
        log(f"CSV exported: {csv_row_count:,} rows | {len(tp.FLAT_FACT_COLUMNS)} columns.")

        # ── Step 6: Read CSV and write XLSX ───────────────────────────────────
        log("Reading CSV for Excel workbook …")
        flat_rows = _read_csv_as_dicts(tmp_csv_path)

        xlsx_name = f"{base_name}.xlsx"
        xlsx_path = out_dir / xlsx_name

        log(f"Writing Excel workbook ({len(flat_rows):,} data rows) …")

        # Collect all log lines generated so far, deduplicate, preserve order
        final_log_lines = list(jobs.get_job(job_id)["log_lines"]) + collected_log_lines
        seen: set[str] = set()
        deduped_log: list[str] = []
        for line in final_log_lines:
            if line not in seen:
                seen.add(line)
                deduped_log.append(line)

        _write_xlsx(xlsx_path, flat_rows, deduped_log)
        log(f"✅ Excel workbook saved → {xlsx_path.resolve()}")

        output_files = [str(xlsx_path.resolve())]

        # ── Step 7: Optional dimension / fact CSV exports ─────────────────────
        if export_dims and masters_global is not None:
            for dim in export_dims:
                dim_key = dim.upper()
                dim_rows = masters_global.dimensions.get(dim_key, [])
                if dim_rows:
                    p = out_dir / f"dim_{dim.lower()}.csv"
                    tp.write_csv_rows(p, dim_rows)
                    output_files.append(str(p.resolve()))
                    log(f"Dimension table written → {p.name}")

        if "voucher" in export_facts:
            # Export raw voucher headers from SQLite via a custom query
            log("Exporting fact_voucher.csv is not directly available in V2 (voucher headers are merged into the flat fact table).")

        if "ledger_entry" in export_facts:
            log("Exporting fact_ledger_entry.csv is not directly available in V2 (rows are in the flat fact table / SQLite).")

        if "inventory_line" in export_facts:
            log("Exporting fact_inventory_line.csv is not directly available in V2 (rows are in the flat fact table / SQLite).")

        log(
            f"✅ Extraction complete! "
            f"{len(flat_rows):,} rows | {len(output_files)} file(s) written."
        )
        jobs.set_output_files(job_id, output_files)
        jobs.set_status(job_id, jobs.COMPLETED)

    except Exception as exc:  # noqa: BLE001
        err_msg = str(exc)
        logger.exception("Extraction job %s failed", job_id)
        jobs.append_log(job_id, f"❌ ERROR: {err_msg}")
        jobs.set_error(job_id, err_msg)
        jobs.set_status(job_id, jobs.FAILED)

    finally:
        root_logger.removeHandler(handler)
        tp_logger.removeHandler(handler)

        # Clean up temp SQLite DB
        if db_path and db_path.exists():
            try:
                db_path.unlink()
            except Exception:
                pass

        # Clean up temp CSV
        if tmp_csv_path and tmp_csv_path.exists():
            try:
                tmp_csv_path.unlink()
            except Exception:
                pass

        # Clean up uploaded XML file
        if xml_file:
            try:
                Path(xml_file).unlink(missing_ok=True)
            except Exception:
                pass


def start_extraction_job(job_id: str, params: dict[str, Any]) -> None:
    """Spawn a background thread to run the extraction and return immediately."""
    t = threading.Thread(
        target=run_extraction,
        args=(job_id, params),
        daemon=True,
        name=f"tally-job-{job_id[:8]}",
    )
    t.start()
